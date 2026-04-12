import csv
import json
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.crm import Customer, Supplier
from app.models.import_batch import ImportBatch
from app.models.order import Order
from app.models.task import Task
from app.models.user import User
from app.services.audit import log_audit, serialize_model
from app.services.orders import refresh_order_total


@dataclass
class ParsedSheet:
    filename: str
    headers: list[str]
    rows: list[dict[str, Any]]


def _decode_csv(contents: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            return contents.decode(encoding)
        except UnicodeDecodeError:
            continue
    return contents.decode("utf-8", errors="replace")


def _coerce_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (date, Decimal, int, float)):
        return value
    return str(value).strip()


async def parse_upload(file: UploadFile, max_rows: int | None = None) -> ParsedSheet:
    contents = await file.read()
    filename = file.filename or "upload"
    suffix = Path(filename).suffix.lower()
    limit = max_rows or get_settings().upload_max_rows

    if suffix == ".csv":
        text = _decode_csv(contents)
        reader = csv.DictReader(StringIO(text))
        headers = reader.fieldnames or []
        rows = [{key: _coerce_cell(value) for key, value in row.items()} for _, row in zip(range(limit), reader)]
        return ParsedSheet(filename=filename, headers=headers, rows=rows)

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows_iter, [])]
        rows = []
        for _, row in zip(range(limit), rows_iter):
            rows.append({headers[index]: _coerce_cell(value) for index, value in enumerate(row) if index < len(headers)})
        return ParsedSheet(filename=filename, headers=headers, rows=rows)

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="نوع الملف غير مدعوم. استخدم CSV أو XLSX.",
    )


def parse_column_map(raw_column_map: str) -> dict[str, str]:
    try:
        column_map = json.loads(raw_column_map)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="خريطة الأعمدة ليست JSON صالحا") from exc
    if not isinstance(column_map, dict):
        raise HTTPException(status_code=400, detail="خريطة الأعمدة يجب أن تكون كائنا")
    return {str(key): str(value) for key, value in column_map.items() if value}


def _mapped(row: dict[str, Any], column_map: dict[str, str]) -> dict[str, Any]:
    return {field: row.get(source) for field, source in column_map.items()}


def _as_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except InvalidOperation as exc:
        raise ValueError(f"قيمة رقمية غير صالحة: {value}") from exc


def _as_int(value: Any, default: int = 1) -> int:
    if value in (None, ""):
        return default
    return int(float(str(value).replace(",", "")))


def build_import_object(target_module: str, data: dict[str, Any]) -> Any:
    if target_module == "customers":
        if not data.get("name"):
            raise ValueError("اسم العميل مطلوب")
        return Customer(**{key: value for key, value in data.items() if key in Customer.__table__.columns})

    if target_module == "suppliers":
        if not data.get("name"):
            raise ValueError("اسم المورد مطلوب")
        return Supplier(**{key: value for key, value in data.items() if key in Supplier.__table__.columns})

    if target_module == "tasks":
        if not data.get("title"):
            raise ValueError("عنوان المهمة مطلوب")
        allowed = {key: value for key, value in data.items() if key in Task.__table__.columns}
        return Task(**allowed)

    if target_module == "orders":
        if not data.get("product_name"):
            raise ValueError("اسم المنتج مطلوب")
        order = Order(
            product_name=data["product_name"],
            customer_id=_as_int(data.get("customer_id"), 0) or None,
            supplier_id=_as_int(data.get("supplier_id"), 0) or None,
            status=data.get("status") or "draft",
            currency=data.get("currency") or "USD",
            quantity=_as_int(data.get("quantity"), 1),
            unit_price=_as_decimal(data.get("unit_price")),
            shipping_fee=_as_decimal(data.get("shipping_fee")),
            notes=data.get("notes"),
        )
        refresh_order_total(order)
        return order

    raise ValueError("الوحدة المستهدفة غير مدعومة للاستيراد")


async def commit_import(
    db: Session,
    *,
    file: UploadFile,
    target_module: str,
    column_map: dict[str, str],
    actor: User,
) -> ImportBatch:
    parsed = await parse_upload(file)
    errors: list[dict[str, Any]] = []
    successful_rows = 0

    batch = ImportBatch(
        filename=parsed.filename,
        target_module=target_module,
        status="processing",
        total_rows=len(parsed.rows),
        column_map=column_map,
        created_by_id=actor.id,
    )
    db.add(batch)
    db.flush()

    for index, row in enumerate(parsed.rows, start=2):
        try:
            with db.begin_nested():
                obj = build_import_object(target_module, _mapped(row, column_map))
                db.add(obj)
                db.flush()
                log_audit(
                    db,
                    actor=actor,
                    action="import_create",
                    entity_type=target_module,
                    entity_id=getattr(obj, "id"),
                    new_values=serialize_model(obj),
                )
            successful_rows += 1
        except Exception as exc:  # noqa: BLE001 - per-row import must collect validation failures.
            errors.append({"row": index, "message": str(exc), "data": row})

    batch.successful_rows = successful_rows
    batch.failed_rows = len(errors)
    batch.errors = errors
    batch.status = "completed_with_errors" if errors else "completed"
    log_audit(
        db,
        actor=actor,
        action="import_commit",
        entity_type="ImportBatch",
        entity_id=batch.id,
        new_values={
            "target_module": target_module,
            "total_rows": batch.total_rows,
            "successful_rows": batch.successful_rows,
            "failed_rows": batch.failed_rows,
        },
    )
    db.commit()
    db.refresh(batch)
    return batch
